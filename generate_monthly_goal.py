#!/usr/bin/env python3
"""
위베이프 월간 목표 설정 엑셀 생성기
사용법: python generate_monthly_goal.py [연도] [월]
예시:   python generate_monthly_goal.py 2026 8
인수 없이 실행하면 기본값(2026년 7월)으로 생성
"""

import sys
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════
#  ▶ 매월 수정 영역 (또는 커맨드라인 인수 사용)
# ══════════════════════════════════════════
DEFAULT_YEAR  = 2026
DEFAULT_MONTH = 7
# ══════════════════════════════════════════

STORES = [
    "인천 연수점",
    "인천 논현점",
    "인천 구월 로데오점",
    "인천 구월 길병원점",
    "부천 상동점",
    "부천 신중동점",
    "인천공항점",
    "검단점",
    "계산점",
]

# ── 고정 행 번호 (모든 매장 시트 동일) ──
ROWS = {
    "prev":    5,   # 전월 실적
    "target":  6,   # 이번 달 목표 매출
    "trend":   9,   # 현재 매출 추이
    "cust":   10,   # 고객 유입 상황
    "top1":   13,   # 주력 제품 1위
    "top2":   14,   # 주력 제품 2위
    "top3":   15,   # 주력 제품 3위
    "staff":  18,   # 직원 현장 의견
    "plan":   19,   # 매출 증대 방안
    "special":20,   # 특이사항
}

# ── 색상 팔레트 ──
C_BRAND     = "1F4E79"   # 진네이비 (헤더)
C_SECTION   = "2E75B6"   # 미디엄블루 (섹션)
C_LABEL     = "D6E4F0"   # 연파랑 (레이블)
C_INPUT_TXT = "FFFDE7"   # 연노랑 (텍스트 입력)
C_INPUT_NUM = "E8F5E9"   # 연초록 (숫자 입력)
C_WHITE     = "FFFFFF"
C_SUMMARY1  = "F0F7FF"   # 요약 홀수행
C_SUMMARY2  = "FFFFFF"   # 요약 짝수행
C_TOTAL     = "EBF3FB"   # 합계행


def _side(s='thin', c='BDBDBD'):
    return Side(style=s, color=c)


def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _outer(ws, r, sc, ec):
    """병합 범위 외곽 테두리 적용"""
    t = _side()
    n = Side(style=None)
    for col in range(sc, ec + 1):
        cell = ws.cell(row=r, column=col)
        cell.border = Border(
            left=t  if col == sc else n,
            right=t if col == ec else n,
            top=t,
            bottom=t,
        )


def _setup_cols(ws):
    ws.column_dimensions['A'].width = 1.5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 8
    ws.sheet_view.showGridLines = False


def _title(ws, r, text):
    ws.merge_cells(f'B{r}:G{r}')
    c = ws.cell(row=r, column=2)
    c.value = text
    c.font = Font(name='맑은 고딕', bold=True, size=15, color=C_WHITE)
    c.fill = PatternFill('solid', fgColor=C_BRAND)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    ws.row_dimensions[r].height = 42


def _subtitle(ws, r, text):
    ws.merge_cells(f'B{r}:G{r}')
    c = ws.cell(row=r, column=2)
    c.value = text
    c.font = Font(name='맑은 고딕', size=9, color='555555')
    c.fill = PatternFill('solid', fgColor='EBF3FB')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    ws.row_dimensions[r].height = 20


def _section(ws, r, text):
    ws.merge_cells(f'B{r}:G{r}')
    c = ws.cell(row=r, column=2)
    c.value = f"  {text}"
    c.font = Font(name='맑은 고딕', bold=True, size=11, color=C_WHITE)
    c.fill = PatternFill('solid', fgColor=C_SECTION)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 28


def _numeric_row(ws, r, label):
    ws.merge_cells(f'B{r}:C{r}')
    lc = ws.cell(row=r, column=2)
    lc.value = label
    lc.font = Font(name='맑은 고딕', bold=True, size=10, color='1F4E79')
    lc.fill = PatternFill('solid', fgColor=C_LABEL)
    lc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    _outer(ws, r, 2, 3)

    ws.merge_cells(f'D{r}:G{r}')
    ic = ws.cell(row=r, column=4)
    ic.fill = PatternFill('solid', fgColor=C_INPUT_NUM)
    ic.number_format = '#,##0"원"'
    ic.alignment = Alignment(horizontal='right', vertical='center', indent=1)
    _outer(ws, r, 4, 7)

    ws.row_dimensions[r].height = 26
    return ic


def _text_row(ws, r, label, sublabel=None, height=50):
    if sublabel:
        lc = ws.cell(row=r, column=2)
        lc.value = label
        lc.font = Font(name='맑은 고딕', bold=True, size=10, color='1F4E79')
        lc.fill = PatternFill('solid', fgColor=C_LABEL)
        lc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        lc.border = _border()

        sc = ws.cell(row=r, column=3)
        sc.value = sublabel
        sc.font = Font(name='맑은 고딕', size=10, color='333333')
        sc.fill = PatternFill('solid', fgColor='E3F0FF')
        sc.alignment = Alignment(horizontal='center', vertical='center')
        sc.border = _border()
    else:
        ws.merge_cells(f'B{r}:C{r}')
        lc = ws.cell(row=r, column=2)
        lc.value = label
        lc.font = Font(name='맑은 고딕', bold=True, size=10, color='1F4E79')
        lc.fill = PatternFill('solid', fgColor=C_LABEL)
        lc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        _outer(ws, r, 2, 3)

    ws.merge_cells(f'D{r}:G{r}')
    ic = ws.cell(row=r, column=4)
    ic.fill = PatternFill('solid', fgColor=C_INPUT_TXT)
    ic.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=1)
    _outer(ws, r, 4, 7)

    ws.row_dimensions[r].height = height
    return ic


def create_store_sheet(wb, store_name, year, month):
    ws = wb.create_sheet(title=store_name)
    _setup_cols(ws)

    r = 1
    _title(ws, r, f"{store_name}  |  {year}년 {month}월 월간 목표 설정")
    # r=1: title

    r = 2
    _subtitle(ws, r, f"  작성 기준: {year}년 {month}월     작성일: ___________     작성자: ___________")
    # r=2: subtitle

    r = 3
    ws.row_dimensions[r].height = 8  # spacer

    # ── ① 매출 현황 ──  (r=4~6)
    r = 4
    _section(ws, r, "① 매출 현황")

    r = 5  # ROWS['prev']
    _numeric_row(ws, r, "전월 실적 (원)")

    r = 6  # ROWS['target']
    _numeric_row(ws, r, "이번 달 목표 매출 (원)")

    r = 7
    ws.row_dimensions[r].height = 6  # spacer

    # ── ② 현황 분석 ──  (r=8~10)
    r = 8
    _section(ws, r, "② 현황 분석")

    r = 9   # ROWS['trend']
    _text_row(ws, r, "현재 매출 추이", height=55)

    r = 10  # ROWS['cust']
    _text_row(ws, r, "고객 유입 상황", height=55)

    r = 11
    ws.row_dimensions[r].height = 6  # spacer

    # ── ③ 주력 제품 TOP 3 ──  (r=12~15)
    r = 12
    _section(ws, r, "③ 주력 제품 TOP 3")

    r = 13  # ROWS['top1']
    _text_row(ws, r, "주력 제품 TOP 3", "1위", height=28)

    r = 14  # ROWS['top2']
    _text_row(ws, r, "", "2위", height=28)

    r = 15  # ROWS['top3']
    _text_row(ws, r, "", "3위", height=28)

    r = 16
    ws.row_dimensions[r].height = 6  # spacer

    # ── ④ 의견 및 전략 ──  (r=17~20)
    r = 17
    _section(ws, r, "④ 의견 및 전략")

    r = 18  # ROWS['staff']
    _text_row(ws, r, "직원 현장 의견", height=65)

    r = 19  # ROWS['plan']
    _text_row(ws, r, "매출 증대 방안", height=65)

    r = 20  # ROWS['special']
    _text_row(ws, r, "특이사항", height=50)

    r = 21
    ws.row_dimensions[r].height = 8  # spacer

    r = 22
    ws.merge_cells(f'B{r}:G{r}')
    fc = ws.cell(row=r, column=2)
    fc.value = "위베이프 (WEVAPE)  —  월간 목표 관리 시스템"
    fc.font = Font(name='맑은 고딕', size=9, color='9E9E9E', italic=True)
    fc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 20


def create_summary_sheet(wb, year, month):
    ws = wb.create_sheet(title="요약", index=0)
    ws.sheet_view.showGridLines = False

    # 컬럼 너비
    col_widths = [4, 18, 14, 14, 10, 26, 17, 17, 17, 26]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    # 제목
    ws.merge_cells(f'A{r}:J{r}')
    c = ws.cell(row=r, column=1)
    c.value = f"위베이프  |  {year}년 {month}월 월간 목표 요약"
    c.font = Font(name='맑은 고딕', bold=True, size=15, color=C_WHITE)
    c.fill = PatternFill('solid', fgColor=C_BRAND)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    ws.row_dimensions[r].height = 42; r += 1

    # 부제
    ws.merge_cells(f'A{r}:J{r}')
    c = ws.cell(row=r, column=1)
    c.value = f"  전체 매장 데이터 자동 취합  |  기준월: {year}년 {month}월  |  매장 시트 입력 시 자동 반영"
    c.font = Font(name='맑은 고딕', size=9, color='555555')
    c.fill = PatternFill('solid', fgColor='EBF3FB')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    ws.row_dimensions[r].height = 20; r += 1

    r += 1  # spacer (row 3)

    # 헤더
    headers = [
        "#", "매장명", "전월 실적", "이번 달 목표",
        "증감률", "현재 매출 추이",
        "주력 제품 1위", "주력 제품 2위", "주력 제품 3위",
        "특이사항"
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col)
        c.value = h
        c.font = Font(name='맑은 고딕', bold=True, size=10, color=C_WHITE)
        c.fill = PatternFill('solid', fgColor=C_BRAND)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = _border()
    ws.row_dimensions[r].height = 32; r += 1

    data_start = r

    for i, store in enumerate(STORES):
        sn = store.replace("'", "''")
        bg = C_SUMMARY1 if i % 2 == 0 else C_SUMMARY2

        def _cell(col, value=None, fmt=None, bold=False, align='left'):
            c = ws.cell(row=r, column=col)
            if value is not None:
                c.value = value
            c.font = Font(name='맑은 고딕', bold=bold, size=10)
            if fmt:
                c.number_format = fmt
            c.fill = PatternFill('solid', fgColor=bg)
            ha = 'center' if align == 'center' else ('right' if align == 'right' else 'left')
            c.alignment = Alignment(horizontal=ha, vertical='top', wrap_text=True, indent=1)
            c.border = _border()
            return c

        _cell(1, i + 1, align='center').font = Font(name='맑은 고딕', size=10, color='777777')
        _cell(2, store, bold=True).font = Font(name='맑은 고딕', bold=True, size=10, color='1F4E79')

        c3 = _cell(3, f"='{sn}'!D{ROWS['prev']}",   fmt='#,##0"원"', align='right')
        c4 = _cell(4, f"='{sn}'!D{ROWS['target']}", fmt='#,##0"원"', align='right')

        # 증감률
        c5 = ws.cell(row=r, column=5)
        c5.value = f"=IFERROR(D{r}/C{r}-1,\"\")"
        c5.font = Font(name='맑은 고딕', size=10)
        c5.number_format = '0.0%'
        c5.fill = PatternFill('solid', fgColor=bg)
        c5.alignment = Alignment(horizontal='center', vertical='center')
        c5.border = _border()

        _cell(6,  f"='{sn}'!D{ROWS['trend']}").font   = Font(name='맑은 고딕', size=9)
        _cell(7,  f"='{sn}'!D{ROWS['top1']}").font    = Font(name='맑은 고딕', size=9)
        _cell(8,  f"='{sn}'!D{ROWS['top2']}").font    = Font(name='맑은 고딕', size=9)
        _cell(9,  f"='{sn}'!D{ROWS['top3']}").font    = Font(name='맑은 고딕', size=9)
        _cell(10, f"='{sn}'!D{ROWS['special']}").font = Font(name='맑은 고딕', size=9)

        ws.row_dimensions[r].height = 50
        r += 1

    # 합계행
    bg = C_TOTAL
    for col in range(1, 11):
        c = ws.cell(row=r, column=col)
        c.fill = PatternFill('solid', fgColor=bg)
        c.border = _border()
        c.font = Font(name='맑은 고딕', bold=True, size=10)

    ws.cell(row=r, column=2).value = "합계 / 평균"
    ws.cell(row=r, column=2).alignment = Alignment(horizontal='center', vertical='center')

    end = data_start + len(STORES) - 1
    c3 = ws.cell(row=r, column=3)
    c3.value = f"=SUM(C{data_start}:C{end})"
    c3.number_format = '#,##0"원"'
    c3.alignment = Alignment(horizontal='right', vertical='center', indent=1)

    c4 = ws.cell(row=r, column=4)
    c4.value = f"=SUM(D{data_start}:D{end})"
    c4.number_format = '#,##0"원"'
    c4.alignment = Alignment(horizontal='right', vertical='center', indent=1)

    c5 = ws.cell(row=r, column=5)
    c5.value = f"=IFERROR(D{r}/C{r}-1,\"\")"
    c5.number_format = '0.0%'
    c5.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[r].height = 28; r += 1

    # 푸터
    r += 1
    ws.merge_cells(f'A{r}:J{r}')
    fc = ws.cell(row=r, column=1)
    fc.value = "위베이프 (WEVAPE)  —  월간 목표 관리 시스템  |  각 매장 시트에 데이터 입력 시 이 요약 시트에 자동 반영됩니다"
    fc.font = Font(name='맑은 고딕', size=9, color='9E9E9E', italic=True)
    fc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 20


def main():
    year  = int(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_YEAR
    month = int(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_MONTH

    if not (1 <= month <= 12):
        print(f"오류: 월은 1~12 사이여야 합니다. 입력값: {month}")
        sys.exit(1)

    filename = f"위베이프_월간목표_{year}년{month}월.xlsx"
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 요약 시트 먼저 (index=0으로 첫 번째 탭)
    create_summary_sheet(wb, year, month)

    # 매장 시트 9개
    for store in STORES:
        create_store_sheet(wb, store, year, month)

    wb.save(output_path)

    print(f"\n  파일 생성 완료: {filename}")
    print(f"  저장 위치: {output_path}")
    print(f"  시트 구성: 요약 1개 + 매장 {len(STORES)}개 = 총 {1 + len(STORES)}개")
    print(f"  기준월: {year}년 {month}월")
    print()
    print(f"  다음 달 사용법: python {os.path.basename(__file__)} {year} {month % 12 + 1 if month < 12 else 1}")
    if month == 12:
        print(f"  (연도 변경 주의: python {os.path.basename(__file__)} {year + 1} 1)")
    print()


if __name__ == "__main__":
    main()
